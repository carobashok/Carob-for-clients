"""
vahan_rto_scraper.py
====================
Full RTO-wise + Vehicle-Category-Group-wise vehicle registration data
scraper for the Vahan 4.0 Dashboard.

CONFIRMED WORKFLOW (from manual inspection):
  Type       = Actual Value
  State      = <target state>
  RTO        = <loop through each individual RTO>
  Y-Axis     = Vehicle Class
  X-Axis     = Vehicle Category Group
  Year Type  = Calendar Year
  Year       = <target year>
  → Click Refresh
  → Table shows Vehicle Class rows x sub-columns (e.g. 2WIC/2WN/2WT/TOTAL)
    for ONE Category Group at a time, selectable via a small dropdown
    top-right of the table (TWO WHEELER / THREE WHEELER / FOUR WHEELER /
    GOODS VEHICLES / etc.)
  → We loop through EVERY category group in that dropdown too, for
    each RTO, to get the complete breakdown.

OUTPUT: Long/tidy format — one row per
  (State, RTO, Year, Category_Group, Vehicle_Class, Sub_Column, Value)

Real selector IDs (confirmed via inspect_vahan_page.py):
  Type            -> #j_idt30
  State           -> #j_idt39
  RTO             -> #selectedRto
  Y-Axis          -> #yaxisVar
  X-Axis          -> #xaxisVar
  Year Type       -> #selectedYearType
  Year            -> #selectedYear
  Category Group  -> small dropdown inside table panel, id discovered
                      dynamically (see find_category_group_dropdown)
  Refresh button  -> text 'Refresh'

USAGE:
    python vahan_rto_scraper.py --no-headless --state Puducherry --year 2026
    python vahan_rto_scraper.py --state "Tamil Nadu" --year 2025 --output tn_full_data
"""

import argparse
import asyncio
import random
import pandas as pd
from playwright.async_api import async_playwright, TimeoutError as PlaywrightTimeout

VAHAN_URL = "https://vahan.parivahan.gov.in/vahan4dashboard/vahan/view/reportview.xhtml"


async def sleep_r(lo=0.6, hi=1.3):
    await asyncio.sleep(random.uniform(lo, hi))


def css_id(raw_id: str) -> str:
    """Escape special characters (like ':') for use in a CSS id selector."""
    return raw_id.replace(":", "\\:")


async def close_any_open_pf_panel(page, timeout=2000):
    """
    PrimeFaces selectOneMenu panels can be left open from a PRIOR
    interaction with a DIFFERENT dropdown (e.g. X-Axis panel still open
    when we try to click Year). An open panel can visually overlap and
    intercept clicks meant for other elements. Press Escape and click
    a neutral spot to ensure nothing is left open before proceeding.
    """
    try:
        open_panel = page.locator("div.ui-selectonemenu-panel:visible").first
        if await open_panel.count() > 0 and await open_panel.is_visible(timeout=timeout):
            await page.keyboard.press("Escape")
            await sleep_r(0.3, 0.5)
    except Exception:
        pass


async def open_pf_dropdown(page, trigger_id: str, timeout=15000):
    await wait_for_block_overlay_gone(page, timeout=10000)
    await close_any_open_pf_panel(page)

    safe_id = css_id(trigger_id)
    trigger = page.locator(f"#{safe_id}")
    await trigger.scroll_into_view_if_needed(timeout=timeout)

    panel = page.locator(f"#{safe_id}_panel")

    # If THIS panel is already visible (stuck open from a previous
    # interaction with the same dropdown), close it first, then click fresh.
    try:
        if await panel.is_visible(timeout=1000):
            await page.keyboard.press("Escape")
            await sleep_r(0.3, 0.5)
    except Exception:
        pass

    await trigger.click(timeout=timeout, force=True)
    await sleep_r(0.25, 0.5)

    try:
        await panel.wait_for(state="visible", timeout=timeout)
    except PlaywrightTimeout:
        # Retry once: click again (sometimes first click only focuses, doesn't open)
        await close_any_open_pf_panel(page)
        await trigger.click(timeout=timeout, force=True)
        await sleep_r(0.3, 0.6)
        await panel.wait_for(state="visible", timeout=timeout)

    return panel


async def get_pf_current_label(page, trigger_id: str, timeout=5000) -> str:
    safe_id = css_id(trigger_id)
    try:
        label_el = page.locator(f"#{safe_id}_label")
        return (await label_el.inner_text(timeout=timeout)).strip()
    except Exception:
        return "(unknown)"


async def select_pf_option(page, trigger_id: str, option_text: str, timeout=15000):
    panel = await open_pf_dropdown(page, trigger_id, timeout)
    item = panel.locator(f"li:has-text('{option_text}')").first
    if await item.count() == 0:
        await page.keyboard.press("Escape")
        raise ValueError(f"Option '{option_text}' not found in #{trigger_id}")
    await item.click(timeout=timeout)
    await sleep_r(0.5, 1.0)


async def select_pf_option_verified(page, trigger_id: str, option_text: str, timeout=15000, retries=2):
    """Select an option and verify the dropdown's label actually updated to match."""
    current_label = "(not attempted)"
    for attempt in range(1, retries + 1):
        await select_pf_option(page, trigger_id, option_text, timeout)
        current_label = await get_pf_current_label(page, trigger_id)
        # option_text might be a partial match (e.g. "BAHOUR" matching "BAHOUR - PY11(...)")
        if option_text.lower() in current_label.lower() or current_label.lower() in option_text.lower():
            return current_label
        print(f"        ⚠ Selection mismatch (attempt {attempt}): wanted '{option_text}', label shows '{current_label}' — retrying")
        await sleep_r(0.5, 1.0)
    return current_label  # return whatever we ended up with after retries exhausted


async def get_pf_options(page, trigger_id: str, timeout=15000) -> list[str]:
    panel = await open_pf_dropdown(page, trigger_id, timeout)
    items = await panel.locator("li").all_inner_texts()
    items = [i.strip() for i in items if i.strip()]
    await page.keyboard.press("Escape")
    await sleep_r(0.25, 0.5)
    return items


async def wait_for_block_overlay_gone(page, timeout=15000):
    """
    PrimeFaces shows a blocking overlay (id often containing 'blockpnl')
    during AJAX calls triggered by Refresh. We wait for any visible
    blocking panel to disappear before proceeding.
    """
    try:
        overlay = page.locator("[id*='blockpnl'], .ui-blockui").first
        if await overlay.count() > 0:
            await overlay.wait_for(state="hidden", timeout=timeout)
    except Exception:
        pass  # No overlay present, or it cleared fast — fine either way


async def click_refresh(page, timeout=10000):
    btn = page.locator("button:has-text('Refresh'), a:has-text('Refresh')").first
    await btn.scroll_into_view_if_needed(timeout=timeout)
    await sleep_r(0.3, 0.5)
    await btn.click(timeout=timeout, force=True)
    await sleep_r(0.3, 0.5)
    await wait_for_block_overlay_gone(page)


CATEGORY_GROUP_DROPDOWN_ID = "vchgroupTable:selectCatgGrp"


async def find_category_group_dropdown_id(page) -> str | None:
    """
    CONFIRMED via HTML inspection: this dropdown has a stable PrimeFaces
    naming-container id 'vchgroupTable:selectCatgGrp' (the colon is normal
    PrimeFaces convention, not a dynamic id). No heuristics needed.
    """
    locator = page.locator(f"#{css_id(CATEGORY_GROUP_DROPDOWN_ID)}")
    try:
        if await locator.count() > 0:
            return CATEGORY_GROUP_DROPDOWN_ID
    except Exception:
        pass
    return None


async def get_category_groups(page, dropdown_id: str) -> list[str]:
    try:
        return await get_pf_options(page, dropdown_id)
    except Exception:
        return []


async def select_category_group(page, dropdown_id: str, group_name: str):
    await select_pf_option(page, dropdown_id, group_name)


async def get_table_title(page) -> str:
    """Read the blue header bar text above the table, e.g.
    'Vehicle Class Wise Vehicle Category Group Data of BAHOUR...(2026)'
    Useful to confirm the table actually updated after a selection change."""
    try:
        title_el = page.locator(".ui-datatable, .ui-panel").locator("text=/Wise.*Data/i").first
        if await title_el.count() > 0:
            return (await title_el.inner_text(timeout=2000)).strip()
    except Exception:
        pass
    return "(title not found)"


def title_matches_expected(title: str, expected_rto_fragment: str, expected_year: str) -> bool:
    """
    CRITICAL DATA-INTEGRITY CHECK. Requires the title to explicitly
    contain BOTH the expected RTO name fragment AND the expected year.
    'For All State' titles always fail (RTO selection didn't take effect).

    Special case: 'Till Today' doesn't appear in the title — the dashboard
    shows the current calendar year instead (e.g. '(2026)'). For this
    case we skip the year check and only validate the RTO fragment.
    """
    if not title or "title not found" in title.lower():
        return False
    if "for all state" in title.lower():
        return False
    title_lower = title.lower()
    rto_fragment_lower = expected_rto_fragment.lower().strip()
    if rto_fragment_lower not in title_lower:
        return False
    # 'Till Today' won't appear in the title — the dashboard shows the
    # current year instead, which we don't know in advance. Skip year check.
    if expected_year == "Till Today":
        return True
    if str(expected_year) not in title:
        return False
    return True


async def wait_for_correct_table_title(page, expected_rto_fragment: str, expected_year: str,
                                        timeout=12000, poll_interval=0.4) -> tuple[bool, str]:
    """
    Polls get_table_title() repeatedly until it matches the expected
    RTO + Year, or timeout elapses. Returns (success, last_title_seen).
    This is the safeguard against the stale-data bug described in
    title_matches_expected's docstring — we refuse to proceed with
    extraction until the title genuinely reflects our current selection.
    """
    elapsed = 0.0
    last_title = ""
    while elapsed < timeout / 1000:
        last_title = await get_table_title(page)
        if title_matches_expected(last_title, expected_rto_fragment, expected_year):
            return True, last_title
        await asyncio.sleep(poll_interval)
        elapsed += poll_interval
    return False, last_title


async def get_table_category_group_header(page) -> str:
    """
    Reads the category-group name from the table's 2nd header row (index 1).
    Table header structure:
      Row 0: 'S No' | 'Maker/Vehicle Class' | 'Vehicle Category Group' (spans N)
      Row 1:  --    |         --             | 'TWO WHEELER' (spans N)  ← we want this
      Row 2:  --    |         --             | '4WIC' | 'LMV' | 'MMV' | 'TOTAL'
    We read the last <th> in row index 1.
    """
    try:
        table = await get_data_table_locator(page)
        thead_rows = table.locator("thead tr")
        n_rows = await thead_rows.count()
        if n_rows < 2:
            return "(category not found)"
        row1_cells = thead_rows.nth(1).locator("th")
        n_cells = await row1_cells.count()
        if n_cells == 0:
            return "(category not found)"
        txt = (await row1_cells.last.inner_text(timeout=2000)).strip()
        return txt if txt else "(category not found)"
    except Exception:
        return "(category not found)"


async def wait_for_category_group_header(page, expected_cg: str,
                                          timeout=10000, poll_interval=0.3) -> tuple[bool, str]:
    """
    Polls get_table_category_group_header() until it matches expected_cg,
    or timeout elapses. Returns (success, last_header_seen).
    """
    elapsed = 0.0
    last_header = ""
    while elapsed < timeout / 1000:
        last_header = await get_table_category_group_header(page)
        if last_header.upper() == expected_cg.upper():
            return True, last_header
        await asyncio.sleep(poll_interval)
        elapsed += poll_interval
    return False, last_header





KNOWN_FILTER_WIDGET_IDS = {"VhCatg", "norms", "fuel", "VhClass"}


async def get_data_table_locator(page):
    """
    CONFIRMED via structural diagnostic dump: there are 7 <table> elements
    on this page. Four are unrelated checkbox-filter widgets with known
    ids (VhCatg, norms, fuel, VhClass) — NOT our data table. The real data
    table has NO id and NO class, and is identified by having a non-empty
    <tbody> with actual data rows (not the multi-row header-only table).

    Strategy: enumerate all <table> elements, skip known filter-widget
    ids, then return the one with tbody rows. If none currently have
    tbody rows (e.g. mid-render), fall back to a positional guess based
    on the confirmed structural dump (6th table, index 5).
    """
    all_tables = page.locator("table")
    count = await all_tables.count()

    candidates = []
    for i in range(count):
        t = all_tables.nth(i)
        tid = await t.get_attribute("id")
        if tid in KNOWN_FILTER_WIDGET_IDS:
            continue
        tr_count = await t.locator("tbody tr").count()
        candidates.append((i, t, tr_count))

    # Prefer a candidate that currently has tbody rows
    for i, t, tr_count in candidates:
        if tr_count > 0:
            return t

    # Fallback: return the candidate matching the confirmed structural
    # position (index 5 in the full table list) if it's among candidates
    for i, t, tr_count in candidates:
        if i == 5:
            return t

    # Last resort: return the last non-filter-widget table found
    if candidates:
        return candidates[-1][1]

    # Absolute last resort: just return table #5 directly even if it
    # wasn't in candidates (shouldn't normally happen)
    return all_tables.nth(5)


async def wait_for_table_rows(page, min_rows=1, timeout=12000):
    """
    After selecting a category group, the table header updates almost
    immediately but tbody rows can lag slightly behind (separate render
    pass). Poll by repeatedly re-evaluating get_data_table_locator
    (which itself looks for a table with tbody rows) until rows appear,
    or timeout elapses.
    """
    elapsed = 0
    step_ms = 300
    while elapsed < timeout:
        table = await get_data_table_locator(page)
        tr_count = await table.locator("tbody tr").count()
        if tr_count >= min_rows:
            await sleep_r(0.2, 0.4)  # tiny settle time
            return
        await asyncio.sleep(step_ms / 1000)
        elapsed += step_ms
    # Timed out — caller will see 0 rows and can log/handle it


async def extract_rows_via_playwright(page) -> tuple[list[str], list[list[str]]]:
    """
    Direct DOM-based extraction using Playwright locators instead of
    pd.read_html on a static HTML snapshot. This sees the live rendered
    DOM, which matters if PrimeFaces lazy-renders/virtualizes rows.
    Returns (header_texts, list_of_row_cell_texts).
    """
    table = await get_data_table_locator(page)

    # Header: last header row (the one with actual sub-column names like 4WIC/LMV/TOTAL)
    header_rows = table.locator("thead tr")
    header_count = await header_rows.count()
    headers = []
    if header_count > 0:
        last_header = header_rows.nth(header_count - 1)
        headers = await last_header.locator("th").all_inner_texts()
        headers = [h.strip() for h in headers]

    # Body rows
    body_rows = table.locator("tbody tr")
    row_count = await body_rows.count()
    print(f"        [pw-extract] thead has {header_count} row(s), tbody has {row_count} row(s)")

    data_rows = []
    for i in range(row_count):
        cells = await body_rows.nth(i).locator("td").all_inner_texts()
        cells = [c.strip() for c in cells]
        if cells:
            data_rows.append(cells)

    return headers, data_rows


def rows_to_long_format(headers: list[str], data_rows: list[list[str]],
                         state, rto, year, category_group, yaxis_label: str = "Vehicle_Class") -> list[dict]:
    """
    Convert Playwright-extracted header+rows directly into long format.
    yaxis_label controls the output column name for the row-category
    field (e.g. 'Vehicle_Class' when Y-Axis=Vehicle Class, or 'Maker'
    when Y-Axis=Maker) — the underlying table structure is identical
    regardless of which Y-Axis dimension was selected on the dashboard.
    """
    rows = []
    if not data_rows:
        return rows

    # Figure out how many leading non-data columns there are (S No, <Y-Axis dimension>)
    # by comparing row length to header length.
    for row_cells in data_rows:
        if len(row_cells) < 2:
            continue
        # Heuristic: first cell is S No (numeric), second is the Y-Axis category (text)
        row_category = row_cells[1] if len(row_cells) > 1 else row_cells[0]
        value_cells = row_cells[2:] if len(row_cells) > 2 else []

        # Align value_cells with headers (skip leading non-numeric header labels if mismatched)
        sub_cols = headers[-len(value_cells):] if len(headers) >= len(value_cells) else \
                   [f"col_{i}" for i in range(len(value_cells))]

        for sub_col, raw_val in zip(sub_cols, value_cells):
            val = pd.to_numeric(str(raw_val).replace(",", ""), errors="coerce")
            rows.append({
                "State": state,
                "RTO": rto,
                "Year": year,
                "Category_Group": category_group,
                yaxis_label: row_category,
                "Sub_Column": sub_col,
                "Value": val,
            })
    return rows


async def extract_current_table(page) -> pd.DataFrame | None:
    table = await get_data_table_locator(page)
    try:
        await table.wait_for(state="attached", timeout=15000)
    except PlaywrightTimeout:
        print("        ⚠ extract_current_table: no table found on page")
        return None
    await sleep_r(0.4, 0.8)

    html = await table.inner_html()
    print(f"        [extract] table inner_html length = {len(html)} chars")

    from io import StringIO
    try:
        dfs = pd.read_html(StringIO(f"<table>{html}</table>"))
    except Exception as e:
        print(f"        ⚠ extract_current_table: read_html failed: {type(e).__name__}: {e}")
        return None
    if not dfs:
        print("        ⚠ extract_current_table: read_html returned 0 tables")
        return None

    print(f"        [extract] read_html found {len(dfs)} sub-table(s); shapes: {[d.shape for d in dfs]}")
    # Pick the dataframe with the most rows (the real data table, not a header/wrapper)
    df = max(dfs, key=lambda d: d.shape[0])

    if df.shape[0] == 0:
        # Dump raw HTML so we can inspect why no rows were captured
        try:
            with open("zero_row_table_dump.html", "w", encoding="utf-8") as f:
                f.write(html)
            print(f"        [extract] 0 rows — dumped raw table HTML to zero_row_table_dump.html ({len(html)} chars)")
        except Exception as e:
            print(f"        [extract] could not write debug dump: {e}")
    if isinstance(df.columns, pd.MultiIndex):
        # Flatten: keep only the lowest meaningful level (sub-column names)
        new_cols = []
        for col in df.columns:
            parts = [str(c) for c in col if "Unnamed" not in str(c)]
            new_cols.append(parts[-1] if parts else str(col[-1]))
        df.columns = new_cols
    df.columns = [str(c).strip() for c in df.columns]
    return df


def melt_table_to_long(df: pd.DataFrame, state, rto, year, category_group,
                        yaxis_label: str = "Vehicle_Class") -> list[dict]:
    """
    Convert a wide table (Y-Axis category rows x sub-columns) into long
    format: one row per (yaxis_label value, Sub_Column, Value).
    Skips 'S No' and identifies the Y-Axis category column automatically.
    yaxis_label controls the output column name (e.g. 'Vehicle_Class' or
    'Maker') depending on which Y-Axis dimension was selected.
    """
    rows = []
    if df is None:
        print("        [melt] df is None")
        return rows
    if df.empty:
        print("        [melt] df is empty (0 rows)")
        return rows

    print(f"        [melt] df shape={df.shape}, columns={list(df.columns)}")

    # Identify the Y-Axis category column (usually 2nd column, text-based)
    class_col = None
    for c in df.columns:
        cl = str(c).lower()
        if "vehicle class" in cl or "class" in cl or "maker" in cl:
            class_col = c
            break
    if class_col is None:
        # Fallback: first non-numeric-looking column after S No
        for c in df.columns:
            if df[c].dtype == object:
                class_col = c
                break
    if class_col is None:
        print(f"        [melt] Could not identify category column among: {list(df.columns)}")
        return rows

    value_cols = [c for c in df.columns if c not in ("S No", class_col)]

    for _, row in df.iterrows():
        row_category = row.get(class_col)
        if pd.isna(row_category):
            continue
        for vc in value_cols:
            raw_val = row.get(vc)
            try:
                val = pd.to_numeric(str(raw_val).replace(",", ""), errors="coerce")
            except Exception:
                val = None
            rows.append({
                "State": state,
                "RTO": rto,
                "Year": year,
                "Category_Group": category_group,
                yaxis_label: str(row_category).strip(),
                "Sub_Column": str(vc).strip(),
                "Value": val,
            })
    return rows


async def find_state_dropdown_id(page) -> str:
    """
    CONFIRMED ISSUE: PrimeFaces auto-generates ids like 'j_idt37'/'j_idt39'
    for unnamed components (Type and State dropdowns), and this numbering
    SHIFTS slightly between sessions/page loads depending on minor
    rendering differences elsewhere on the page. Named ids (selectedRto,
    yaxisVar, xaxisVar, etc.) stay stable — only these generic ones drift.

    Reliable detection: the State dropdown is consistently the 2nd
    ui-selectonemenu div on the page (1st = Type/"Actual Value" style
    selector, 2nd = State). We confirm by checking its default label
    looks like a state-list label (contains 'Running States').
    """
    menus = await page.locator("div.ui-selectonemenu").all()
    for m in menus:
        mid = await m.get_attribute("id")
        if not mid:
            continue
        try:
            label = await m.locator(".ui-selectonemenu-label").inner_text(timeout=1500)
        except Exception:
            continue
        if "running states" in label.lower():
            return mid

    # Fallback: assume it's the 2nd selectOneMenu by position
    if len(menus) >= 2:
        mid = await menus[1].get_attribute("id")
        if mid:
            return mid

    raise RuntimeError("Could not locate the State dropdown by any method")


async def scrape_full_breakdown_multi_year(state: str, years: list[str], headless: bool,
                                            yaxis_choice: str = "Vehicle Class",
                                            xaxis_choice: str = "Vehicle Category Group") -> pd.DataFrame:
    async with async_playwright() as pw:
        browser = await pw.chromium.launch(headless=headless, slow_mo=30)
        ctx = await browser.new_context(
            viewport={"width": 1500, "height": 950},
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
            )
        )
        page = await ctx.new_page()

        print("[setup] Opening Vahan Dashboard …")
        await page.goto(VAHAN_URL, wait_until="networkidle", timeout=60_000)
        await sleep_r(2, 3)

        print("[setup] Detecting State dropdown id (PrimeFaces auto-ids can shift between sessions) …")
        state_dropdown_id = await find_state_dropdown_id(page)
        print(f"        State dropdown id this session: '{state_dropdown_id}'")

        print(f"[setup] Selecting state: {state}")
        all_states = await get_pf_options(page, state_dropdown_id)
        state_match = next((s for s in all_states if state.lower() in s.lower()), None)
        if not state_match:
            raise ValueError(f"State '{state}' not found. Sample available: {all_states[:10]}")
        confirmed_state = await select_pf_option_verified(page, state_dropdown_id, state_match)
        print(f"        State dropdown now shows: '{confirmed_state}'")
        await sleep_r(2, 3)

        print(f"[setup] Ensuring Y-Axis='{yaxis_choice}', X-Axis='{xaxis_choice}' …")
        try:
            await select_pf_option(page, "yaxisVar", yaxis_choice)
            await sleep_r(0.6, 1.0)
        except Exception as e:
            print(f"        ⚠ Y-Axis selection issue: {e}")
        try:
            await select_pf_option(page, "xaxisVar", xaxis_choice)
            await sleep_r(0.6, 1.0)
        except Exception as e:
            print(f"        ⚠ X-Axis selection issue: {e}")

        # Output column name for the Y-Axis dimension's row-category values,
        # e.g. 'Vehicle_Class' when yaxis_choice='Vehicle Class', or 'Maker'
        # when yaxis_choice='Maker'. Keeps the CSV column name meaningful
        # regardless of which Y-Axis dimension was scraped.
        yaxis_column_label = yaxis_choice.replace(" ", "_")

        all_rows = []
        total_years = len(years)

        for yi, year in enumerate(years, start=1):
            print(f"\n{'='*60}")
            print(f"[YEAR {yi}/{total_years}] {year}")
            print(f"{'='*60}")

            print(f"[setup] Selecting year: {year}")
            try:
                confirmed_year = await select_pf_option_verified(page, "selectedYear", year)
                print(f"        Year dropdown now shows: '{confirmed_year}'")
            except Exception as e:
                print(f"        ⚠ Could not select year '{year}': {e} — skipping this year")
                continue
            await sleep_r(1, 2)

            # Re-read RTO list fresh for THIS year — labels can include
            # registration/start dates that may shift, or the list itself
            # may differ slightly depending on year context. Reusing a
            # list captured for a different year caused 'option not found'
            # errors on every RTO.
            print(f"[setup] Reading RTO list for {year} …")
            rto_options = await get_pf_options(page, "selectedRto")
            individual_rtos = [r for r in rto_options if not r.lower().startswith("all vahan4")]
            print(f"        Found {len(individual_rtos)} individual RTOs for {year}")
            total_rtos = len(individual_rtos)

            for i, rto_name in enumerate(individual_rtos, start=1):
                print(f"\n[{year}][{i}/{total_rtos}] RTO: {rto_name}")
                try:
                    confirmed_label = await select_pf_option_verified(page, "selectedRto", rto_name)
                    print(f"        RTO dropdown now shows: '{confirmed_label}'")
                    await sleep_r(0.6, 1.2)

                    await click_refresh(page)
                    await sleep_r(1.0, 1.5)

                    rto_name_fragment = rto_name.split("-")[0].strip()

                    # Wait for the table title to confirm THIS RTO's data has loaded.
                    # If it doesn't transition within the first timeout, retry once:
                    # re-select the RTO and click Refresh again. This handles cases
                    # where a prior RTO's AJAX response arrives late and overwrites the
                    # selection (seen as THIRUBUVANAI still showing when YANAM expected).
                    print(f"        Waiting for table to confirm RTO '{rto_name_fragment}' / Year '{year}' …")
                    rto_title_ok, rto_title = await wait_for_correct_table_title(
                        page, rto_name_fragment, year, timeout=15000
                    )

                    if not rto_title_ok:
                        print(f"        ⚠ Title not matching yet (showing '{rto_title}') — retrying RTO selection + Refresh …")
                        await close_any_open_pf_panel(page)
                        await select_pf_option_verified(page, "selectedRto", rto_name)
                        await sleep_r(1.0, 1.5)
                        await click_refresh(page)
                        await sleep_r(2.0, 3.0)
                        rto_title_ok, rto_title = await wait_for_correct_table_title(
                            page, rto_name_fragment, year, timeout=20000
                        )

                    if not rto_title_ok:
                        print(f"        ❌ RTO-level SKIPPED entirely — table title '{rto_title}' "
                              f"never showed expected RTO '{rto_name_fragment}' / Year '{year}' "
                              f"even after retry. This RTO will be missing from the output.")
                        continue
                    print(f"        ✅ Table confirmed: '{rto_title}'")

                    cg_dropdown_id = await find_category_group_dropdown_id(page)
                    if not cg_dropdown_id:
                        print("        ⚠ Could not locate Category Group dropdown — extracting single visible table only")
                        title_ok, title = await wait_for_correct_table_title(page, rto_name_fragment, year, timeout=12000)
                        if not title_ok:
                            print(f"        ❌ SKIPPED — table title '{title}' never matched expected "
                                  f"RTO '{rto_name_fragment}' / Year '{year}'")
                            continue
                        await wait_for_table_rows(page, timeout=12000)
                        df = await extract_current_table(page)
                        rows = melt_table_to_long(df, state_match, rto_name, year, "UNKNOWN", yaxis_column_label)
                        all_rows.extend(rows)
                        print(f"        → {len(rows)} data points extracted")
                        continue

                    category_groups = await get_category_groups(page, cg_dropdown_id)
                    if not category_groups:
                        print("        ⚠ 0 category groups found, retrying after short wait …")
                        await sleep_r(1.5, 2.0)
                        cg_dropdown_id = await find_category_group_dropdown_id(page)
                        if cg_dropdown_id:
                            category_groups = await get_category_groups(page, cg_dropdown_id)

                    print(f"        Found {len(category_groups)} category groups: {category_groups}")

                    if not category_groups:
                        print("        ⚠ Still 0 category groups — extracting single visible table as fallback")
                        title_ok, title = await wait_for_correct_table_title(page, rto_name_fragment, year, timeout=12000)
                        if not title_ok:
                            print(f"        ❌ SKIPPED — table title '{title}' never matched expected "
                                  f"RTO '{rto_name_fragment}' / Year '{year}'")
                            continue
                        await wait_for_table_rows(page, timeout=12000)
                        df = await extract_current_table(page)
                        rows = melt_table_to_long(df, state_match, rto_name, year, "UNKNOWN", yaxis_column_label)
                        all_rows.extend(rows)
                        print(f"        → {len(rows)} data points extracted")
                        continue

                    # Extract a stable fragment of the RTO name to verify against
                    # the table title (e.g. "BAHOUR" from "BAHOUR - PY11(...)").
                    # This is the core safeguard against the stale-data bug where
                    # a leftover 'All States' default view got mislabeled as
                    # belonging to the currently-selected RTO/Year.
                    rto_name_fragment = rto_name.split("-")[0].strip()

                    for cg in category_groups:
                        try:
                            await select_category_group(page, cg_dropdown_id, cg)
                            await sleep_r(0.4, 0.7)

                            # Gate 1: outer title must show correct RTO + Year
                            title_ok, title = await wait_for_correct_table_title(
                                page, rto_name_fragment, year, timeout=12000
                            )
                            print(f"        [{cg}] title: '{title}' (valid={title_ok})")

                            if not title_ok:
                                print(f"        ❌ [{cg}] SKIPPED — title never matched "
                                      f"RTO '{rto_name_fragment}' / Year '{year}'.")
                                continue

                            # Gate 2: internal table category-group header must match
                            # the category group we just selected. This catches the bug
                            # where Refresh resets to FOUR WHEELER as default, and the
                            # table still shows FOUR WHEELER data while we iterate
                            # through TWO WHEELER / THREE WHEELER etc.
                            cg_ok, cg_header = await wait_for_category_group_header(
                                page, cg, timeout=10000
                            )
                            print(f"        [{cg}] internal header: '{cg_header}' (valid={cg_ok})")

                            if not cg_ok:
                                print(f"        ❌ [{cg}] SKIPPED — internal table header "
                                      f"shows '{cg_header}', expected '{cg}'. "
                                      f"This prevents mislabeled category-group data.")
                                continue

                            await wait_for_table_rows(page, timeout=12000)

                            headers, data_rows = await extract_rows_via_playwright(page)
                            rows = rows_to_long_format(headers, data_rows, state_match, rto_name, year, cg, yaxis_column_label)

                            if not rows:
                                print(f"        [{cg}] Playwright extraction got 0 rows — trying pandas fallback")
                                df = await extract_current_table(page)
                                rows = melt_table_to_long(df, state_match, rto_name, year, cg, yaxis_column_label)

                            all_rows.extend(rows)
                            print(f"        [{cg}] → {len(rows)} data points")
                        except Exception as e:
                            print(f"        ⚠ Category group '{cg}' failed: {type(e).__name__}: {e}")
                            continue

                except Exception as e:
                    print(f"        ❌ Error on RTO '{rto_name}' for year {year}: {type(e).__name__}: {e}")
                    continue

            # Periodic save after each year completes, so a crash later doesn't lose everything
            partial_df = pd.DataFrame(all_rows)
            partial_path = f"_partial_after_{year}.csv"
            try:
                partial_df.to_csv(partial_path, index=False)
                print(f"\n[checkpoint] Saved partial progress through {year} → {partial_path} ({len(partial_df)} rows so far)")
            except Exception as e:
                print(f"[checkpoint] Could not save partial checkpoint: {e}")

        await browser.close()
        return pd.DataFrame(all_rows)


def save_outputs(df: pd.DataFrame, prefix: str):
    if df.empty:
        print("\n❌ No data collected.")
        return

    csv_path = f"{prefix}.csv"
    df.to_csv(csv_path, index=False)
    print(f"\n✅ CSV saved → {csv_path}")

    xlsx_path = f"{prefix}.xlsx"
    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="RTO Data Long")
        ws = writer.sheets["RTO Data Long"]
        from openpyxl.styles import PatternFill, Font, Alignment
        for cell in ws[1]:
            cell.fill = PatternFill("solid", fgColor="0D1B2A")
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center")
        for col in ws.columns:
            w = max(len(str(c.value or "")) for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(w + 4, 40)
    print(f"✅ Excel saved → {xlsx_path}")

    print(f"\nTotal rows (long format): {len(df)}")
    print(f"RTOs covered: {df['RTO'].nunique()}")
    print(f"Category groups covered: {df['Category_Group'].nunique()}")
    print("\nSample rows:")
    print(df.head(15).to_string(index=False))


def parse_args():
    p = argparse.ArgumentParser()
    p.add_argument("--state", default="Puducherry")
    p.add_argument("--start-year", type=int, default=2021)
    p.add_argument("--end-year", type=int, default=2026)
    p.add_argument("--output", default="rto_full_data")
    p.add_argument("--yaxis", default="Vehicle Class",
                   help="Y-Axis dimension on the dashboard, e.g. 'Vehicle Class' or 'Maker'. "
                        "Default 'Vehicle Class' matches original behavior — pass "
                        "--yaxis Maker to scrape Maker-wise data instead.")
    p.add_argument("--xaxis", default="Vehicle Category Group",
                   help="X-Axis dimension on the dashboard, e.g. 'Vehicle Category Group' or 'Fuel'.")
    p.add_argument("--till-today", action="store_true", default=False,
                   help="Also scrape the 'Till Today' cumulative total (all-years combined) "
                        "in addition to the year range. Stored with year='Till Today' in output.")
    p.add_argument("--only-till-today", action="store_true", default=False,
                   help="Scrape ONLY 'Till Today' (skip the individual year range entirely). "
                        "Useful when you want the cumulative stock snapshot only.")
    p.add_argument("--no-headless", dest="headless", action="store_false")
    p.set_defaults(headless=True)
    return p.parse_args()


async def main():
    args = parse_args()

    if args.only_till_today:
        years = ["Till Today"]
    else:
        years = [str(y) for y in range(args.start_year, args.end_year + 1)]
        if args.till_today:
            years = ["Till Today"] + years  # scrape cumulative first, then yearly

    print("=" * 60)
    print("  Vahan Dashboard — Full RTO x Category Group x Year Scraper")
    print(f"  State  : {args.state}")
    print(f"  Years  : {', '.join(years)}")
    print(f"  Y-Axis : {args.yaxis}")
    print(f"  X-Axis : {args.xaxis}")
    print(f"  Mode   : {'Headless' if args.headless else 'Visible browser'}")
    print("=" * 60)

    df = await scrape_full_breakdown_multi_year(
        args.state, years, args.headless,
        yaxis_choice=args.yaxis, xaxis_choice=args.xaxis
    )
    save_outputs(df, args.output)


if __name__ == "__main__":
    asyncio.run(main())
