"""
Carob Technologies — Quote Request Extractor
============================================
Streamlit app: Fetch Gmail → Select emails → Extract with Claude → Store in Supabase

Install:
    pip install streamlit google-auth google-auth-oauthlib
                google-auth-httplib2 google-api-python-client
                anthropic supabase openpyxl

Secrets (.streamlit/secrets.toml):
    ANTHROPIC_API_KEY = "sk-ant-..."

    [supabase]
    url = "https://xxx.supabase.co"
    key = "eyJ..."

    [gmail]
    client_id     = "....apps.googleusercontent.com"
    client_secret = "GOCSPX-..."
    refresh_token = "1//..."
"""

import json
import base64
import re
import os
from datetime import datetime, timezone

import streamlit as st
from google.oauth2.credentials import Credentials
from google.auth.transport.requests import Request
from googleapiclient.discovery import build
import anthropic
from supabase import create_client, Client
from googleapiclient.http import MediaIoBaseUpload
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ── Constants ──────────────────────────────────────────────────────────────────

SCOPES = [
    "https://www.googleapis.com/auth/gmail.readonly",
    "https://www.googleapis.com/auth/gmail.modify",
    "https://www.googleapis.com/auth/drive",
]


EXTRACTION_PROMPT = """You are a quote request extractor for {app_name}.

Extract information from the email below and return ONLY a valid JSON object — no explanation, no markdown, no extra text, no code fences.

Fields to extract:
- customer_name: full name of the sender (string or null)
- customer_email: email address of the original sender (string or null)
- company_name: company or organisation name if mentioned (string or null)
- phone: phone number if mentioned (string or null)
- address: full postal address if mentioned — include door number, street, area, city, pincode (string or null)
- gst_number: GST registration number if mentioned — typically starts with state code digits (string or null)
- product_description: what product or service they are asking about — summarise clearly (string)
- quantity: number of units, licences, projects, etc. (string or null)
- unit: unit type e.g. units, licences, nos, projects (string or null)
- deadline: when they need it by — use their exact words (string or null)
- location: city, state, or project location if mentioned (string or null)
- urgency_level: "high" if urgent/ASAP/immediately, "low" if no timeline, "medium" otherwise (must be: high, medium, or low)
- needs_review: true if ambiguous, incomplete, spam, or unclear (boolean)
- notes: any other relevant context (string or null)

Email subject: {subject}
From: {sender}
Email body:
{body}"""

# Drive folder and schema read from secrets at runtime

STATUS_OPTIONS = ["new", "quoted", "won", "lost", "not_relevant", "not_feasible"]
STATUS_COLORS  = {"new": "🔵", "quoted": "🟡", "won": "🟢", "lost": "🔴", "not_relevant": "⚫", "not_feasible": "🟤"}
URGENCY_ICONS  = {"high": "🔴", "medium": "🟡", "low": "🟢"}


# ── Supabase ───────────────────────────────────────────────────────────────────

@st.cache_resource
def get_supabase() -> Client:
    url = st.secrets["supabase"]["url"]
    key = st.secrets["supabase"]["key"]
    return create_client(url, key)


def get_schema() -> str:
    """Return schema name from secrets. Defaults to 'public'."""
    try:
        return st.secrets["SCHEMA"]
    except Exception:
        return "public"


def get_drive_folder_id() -> str:
    """Return Google Drive parent folder ID from secrets."""
    try:
        return st.secrets["GDRIVE_FOLDER_ID"]
    except Exception:
        return ""


def get_app_name() -> str:
    """Return app display name from secrets."""
    try:
        return st.secrets["APP_NAME"]
    except Exception:
        return "Carob Technologies"


def get_email_provider() -> str:
    """Return email provider from secrets. Must be explicitly set. Supports: gmail, o365"""
    if "EMAIL_PROVIDER" not in st.secrets:
        st.error("EMAIL_PROVIDER is not set in secrets. Please add EMAIL_PROVIDER = \"gmail\" to your secrets.toml.")
        st.stop()
    return st.secrets["EMAIL_PROVIDER"].lower()


def get_followup_tracker_id() -> str:
    """Return Google Drive file ID for Followup_Tracker.xlsx."""
    try:
        return st.secrets["FOLLOWUP_TRACKER_ID"]
    except Exception:
        return ""


# ── Followup Tracker ──────────────────────────────────────────────────────────

def download_tracker_bytes(drive_service) -> bytes | None:
    """Download Followup_Tracker.xlsx from Drive as bytes."""
    tracker_id = get_followup_tracker_id()
    if not tracker_id:
        return None
    try:
        request = drive_service.files().get_media(fileId=tracker_id)
        buf = io.BytesIO()
        from googleapiclient.http import MediaIoBaseDownload
        downloader = MediaIoBaseDownload(buf, request)
        done = False
        while not done:
            _, done = downloader.next_chunk()
        buf.seek(0)
        return buf.read()
    except Exception as e:
        st.warning(f"Could not download tracker: {e}")
        return None


def upload_tracker_bytes(drive_service, data: bytes):
    """Overwrite Followup_Tracker.xlsx in Drive with new bytes."""
    tracker_id = get_followup_tracker_id()
    if not tracker_id:
        return
    try:
        media = MediaIoBaseUpload(
            io.BytesIO(data),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            resumable=False
        )
        drive_service.files().update(fileId=tracker_id, media_body=media).execute()
    except Exception as e:
        st.warning(f"Could not update tracker: {e}")


def append_to_tracker(drive_service, row: dict):
    """Append a new row to Followup_Tracker.xlsx."""
    data = download_tracker_bytes(drive_service)
    if not data:
        return
    wb  = openpyxl.load_workbook(io.BytesIO(data))
    ws  = wb.active
    from openpyxl.styles import Font as XLFont, Alignment as XLAlignment
    link_font = XLFont(name="Arial", color="0563C1", underline="single", size=10)
    normal    = XLFont(name="Arial", size=10)
    left      = XLAlignment(horizontal="left", vertical="center", wrap_text=True)

    next_row = ws.max_row + 1
    ws.cell(next_row, 1,  value=row.get("date", "")).font       = normal
    ws.cell(next_row, 2,  value=row.get("customer_name", "")).font   = normal
    ws.cell(next_row, 3,  value=row.get("company_name", "")).font    = normal
    ws.cell(next_row, 4,  value=row.get("phone", "")).font           = normal
    ws.cell(next_row, 5,  value=row.get("customer_email", "")).font  = normal
    ws.cell(next_row, 6,  value=row.get("product_description", "")).font = normal
    ws.cell(next_row, 7,  value=row.get("status", "quoted")).font    = normal
    ws.cell(next_row, 8,  value="pending").font                      = normal
    ws.cell(next_row, 9,  value="").font                             = normal
    ws.cell(next_row, 10, value="").font                             = normal

    # Drive link as clickable hyperlink
    drive_link = row.get("attachment_folder", "")
    if drive_link:
        link_cell = ws.cell(next_row, 11, value="Open Folder")
        link_cell.hyperlink = drive_link
        link_cell.font      = link_font
    else:
        ws.cell(next_row, 11, value="").font = normal

    # Store quote ID in hidden column L for row removal later
    ws.cell(next_row, 12, value=row.get("id", "")).font = normal

    buf = io.BytesIO()
    wb.save(buf)
    upload_tracker_bytes(drive_service, buf.getvalue())


def remove_from_tracker(drive_service, quote_id: str):
    """Remove a row from Followup_Tracker.xlsx by quote ID (column L)."""
    data = download_tracker_bytes(drive_service)
    if not data:
        return
    wb = openpyxl.load_workbook(io.BytesIO(data))
    ws = wb.active
    rows_to_delete = []
    for row in ws.iter_rows(min_row=2):
        if str(row[11].value) == str(quote_id):
            rows_to_delete.append(row[0].row)
    for r in reversed(rows_to_delete):
        ws.delete_rows(r)
    buf = io.BytesIO()
    wb.save(buf)
    upload_tracker_bytes(drive_service, buf.getvalue())


# ── Merge Quotes ──────────────────────────────────────────────────────────────

def merge_quotes(supabase: Client, source_id: str, target_id: str, service) -> bool:
    """
    Merge source quote into target quote.
    - Appends source conversation_log to target
    - Moves attachments to target Drive folder
    - Deletes source record from Supabase
    - Removes source from Followup Tracker
    """
    try:
        # Fetch both records
        source = supabase.schema(get_schema()).table("quote_requests").select("*").eq("id", source_id).execute().data
        target = supabase.schema(get_schema()).table("quote_requests").select("*").eq("id", target_id).execute().data

        if not source or not target:
            return False

        source = source[0]
        target = target[0]

        # Merge conversation logs
        source_log = source.get("conversation_log") or []
        target_log = target.get("conversation_log") or []

        # Tag merged entries
        for entry in source_log:
            entry["merged_from"] = source.get("customer_email", "")

        merged_log  = target_log + source_log
        reply_count = (target.get("reply_count") or 0) + len(source_log)

        # Update target with merged log
        supabase.schema(get_schema()).table("quote_requests").update({
            "conversation_log": merged_log,
            "reply_count":      reply_count,
            "last_reply_at":    datetime.now(timezone.utc).isoformat(),
            "needs_review":     True,  # flag for review after merge
        }).eq("id", target_id).execute()

        # Move attachments to target Drive folder if both have folders
        source_folder = source.get("attachment_folder", "")
        target_folder = target.get("attachment_folder", "")
        if source_folder and target_folder:
            try:
                drive_service   = get_drive_service()
                source_folder_id = source_folder.split("/")[-1]
                target_folder_id = target_folder.split("/")[-1]
                # List files in source folder
                files = drive_service.files().list(
                    q=f"'{source_folder_id}' in parents",
                    fields="files(id, name)"
                ).execute().get("files", [])
                # Move each file to target folder
                for f in files:
                    drive_service.files().update(
                        fileId=f["id"],
                        addParents=target_folder_id,
                        removeParents=source_folder_id,
                        fields="id"
                    ).execute()
            except Exception:
                pass

        # Remove source from Followup Tracker
        try:
            remove_from_tracker(get_drive_service(), source_id)
        except Exception:
            pass

        # Delete source record from Supabase
        supabase.schema(get_schema()).table("quote_requests").delete().eq("id", source_id).execute()

        return True

    except Exception as e:
        st.error(f"Merge failed: {e}")
        return False


# ── Google Drive ──────────────────────────────────────────────────────────────

@st.cache_resource
def get_drive_service():
    """Build Drive service using same Gmail OAuth credentials (Drive scope included)."""
    gmail_secret = dict(st.secrets["gmail"])
    creds = Credentials(
        token=None,
        refresh_token=gmail_secret["refresh_token"],
        token_uri="https://oauth2.googleapis.com/token",
        client_id=gmail_secret["client_id"],
        client_secret=gmail_secret["client_secret"],
        scopes=SCOPES,
    )
    if not creds.valid:
        creds.refresh(Request())
    return build("drive", "v3", credentials=creds)


def create_drive_folder(drive_service, folder_name: str) -> tuple[str, str]:
    """Create subfolder inside ALIND QUOTES. Returns (folder_id, folder_url)."""
    meta = {
        "name":     folder_name,
        "mimeType": "application/vnd.google-apps.folder",
        "parents":  [get_drive_folder_id()],
    }
    folder = drive_service.files().create(body=meta, fields="id").execute()
    folder_id  = folder.get("id")
    folder_url = f"https://drive.google.com/drive/folders/{folder_id}"
    return folder_id, folder_url


def upload_bytes_to_drive(drive_service, folder_id: str, filename: str, data: bytes, mime_type: str, timestamp: bool = False):
    """Upload raw bytes to a Drive folder. Adds timestamp to filename if timestamp=True."""
    if timestamp:
        name, ext = filename.rsplit(".", 1) if "." in filename else (filename, "")
        ts       = datetime.now().strftime("%Y-%m-%d_%H%M%S")
        filename = f"{name}_{ts}.{ext}" if ext else f"{name}_{ts}"
    meta  = {"name": filename, "parents": [folder_id]}
    media = MediaIoBaseUpload(io.BytesIO(data), mimetype=mime_type or "application/octet-stream")
    drive_service.files().create(body=meta, media_body=media, fields="id").execute()


def generate_quote_excel(email: dict, fields: dict, folder_url: str = "") -> bytes:
    """
    Fill Qtn_table1 right side using zip/XML approach.
    Preserves all comments, formatting and metadata from original template.
    Falls back to simple summary if no template found.
    """
    import zipfile as _zf
    import glob as _glob
    import os as _os

    # Find template file recursively
    xltx_files = _glob.glob("*.xltx") + _glob.glob("**/*.xltx", recursive=True)
    xltx_files  = [_os.path.abspath(f) for f in xltx_files]

    if not xltx_files:
        # Fallback summary
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Quote Request"
        ws["A1"] = "QUOTE REQUEST SUMMARY"
        ws["A2"] = f"Generated: {datetime.now().strftime('%d %b %Y, %I:%M %p')}"
        rows = [
            ("Customer Name", fields.get("customer_name")),
            ("Email",         fields.get("customer_email")),
            ("Company",       fields.get("company_name")),
            ("Phone",         fields.get("phone")),
            ("Address",       fields.get("address")),
            ("GST Number",    fields.get("gst_number")),
            ("Product",       fields.get("product_description")),
            ("Quantity",      f"{fields.get('quantity') or ''} {fields.get('unit') or ''}".strip()),
            ("Deadline",      fields.get("deadline")),
            ("Urgency",       fields.get("urgency_level")),
            ("Notes",         fields.get("notes")),
        ]
        for i, (label, value) in enumerate(rows, start=4):
            ws[f"A{i}"] = label
            ws[f"B{i}"] = value or "—"
        buf = io.BytesIO()
        wb.save(buf)
        return buf.read()

    template_bytes = open(xltx_files[0], "rb").read()

    # Prepare values
    cust_name  = fields.get("customer_name") or ""
    cust_email = fields.get("customer_email") or ""
    company    = fields.get("company_name") or cust_name
    phone      = fields.get("phone") or ""
    gst        = fields.get("gst_number") or ""
    address    = fields.get("address") or fields.get("location") or ""
    addr_parts = ", ".join([p.strip() for p in address.split(",") if p.strip()])

    def esc(s):
        return str(s).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;").replace('"', "&quot;")

    def inject(xml_str, cell_ref, value):
        """Inject text value into cell using inlineStr format to avoid shared string lookup."""
        val = esc(value)
        # Self-closing or existing cell — replace entirely with inlineStr
        m = re.search(rf'<c r="{re.escape(cell_ref)}"([^>]*)/>', xml_str)
        if m:
            attrs = re.sub(r' t="[^"]*"', "", m.group(1))
            new_cell = f'<c r="{cell_ref}"{attrs} t="inlineStr"><is><t>{val}</t></is></c>'
            return xml_str[:m.start()] + new_cell + xml_str[m.end():]
        # Cell with existing content
        m = re.search(rf'<c r="{re.escape(cell_ref)}"([^>]*)>(.*?)</c>', xml_str, re.DOTALL)
        if m:
            attrs = re.sub(r' t="[^"]*"', "", m.group(1))
            new_cell = f'<c r="{cell_ref}"{attrs} t="inlineStr"><is><t>{val}</t></is></c>'
            return xml_str[:m.start()] + new_cell + xml_str[m.end():]
        # Cell not found — insert into row
        row_num = re.match(r"[A-Za-z]+(\d+)", cell_ref).group(1)
        rm = re.search(rf'(<row r="{row_num}"[^>]*>)(.*?)(</row>)', xml_str, re.DOTALL)
        if rm:
            new_c = f'<c r="{cell_ref}" t="inlineStr"><is><t>{val}</t></is></c>'
            return xml_str[:rm.start()] + rm.group(1) + rm.group(2) + new_c + rm.group(3) + xml_str[rm.end():]
        return xml_str

    out_buf = io.BytesIO()
    with _zf.ZipFile(io.BytesIO(template_bytes), "r") as zin:
        with _zf.ZipFile(out_buf, "w", _zf.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                data = zin.read(item.filename)

                if item.filename == "xl/worksheets/sheet2.xml":
                    xml = data.decode("utf-8")
                    xml = inject(xml, "K5",  datetime.now().strftime("%d-%b-%Y"))  # Mail Date
                    if cust_email: xml = inject(xml, "K7",  cust_email)   # Mail ID (below K6 label)
                    if company:    xml = inject(xml, "K9",  company)      # Company Name (below K8 label)
                    if addr_parts: xml = inject(xml, "K11", addr_parts)   # Address (K11:O14 merged)
                    if gst:        xml = inject(xml, "Q14", gst)          # GST# (next to P14 label)
                    if cust_name:  xml = inject(xml, "K16", cust_name)    # Contact Person (below K15 label)
                    if phone:      xml = inject(xml, "K18", phone)        # Phone Numbers (below K17 label)
                    data = xml.encode("utf-8")

                elif item.filename == "xl/drawings/vmlDrawing1.vml":
                    xml = data.decode("utf-8")
                    xml = xml.replace("height:325.5pt", "height:500pt")
                    xml = xml.replace("width:141.5pt",  "width:200pt")
                    data = xml.encode("utf-8")

                elif item.filename == "[Content_Types].xml":
                    xml = data.decode("utf-8")
                    xml = xml.replace(
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.template.main+xml",
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"
                    )
                    data = xml.encode("utf-8")

                elif item.filename == "xl/workbook.xml":
                    xml = data.decode("utf-8")
                    xml = xml.replace(' type="template"', ' type="workbook"')
                    data = xml.encode("utf-8")

                zout.writestr(item, data)

    out_buf.seek(0)
    return out_buf.read()


def create_attachment_subfolder(drive_service, parent_folder_id: str, direction: str) -> str:
    """Create a dated subfolder for attachments. direction: incoming or outgoing."""
    dt_str      = datetime.now().strftime("%Y-%m-%d_%H%M")
    folder_name = f"{dt_str}_{direction}"
    meta = {
        "name":     folder_name,
        "mimeType": "application/vnd.google-apps.folder",
        "parents":  [parent_folder_id],
    }
    folder = drive_service.files().create(body=meta, fields="id").execute()
    return folder.get("id")


def save_to_drive(service, email: dict, fields: dict) -> tuple[str, int]:
    """Create Drive folder, upload attachments in subfolder + Excel. Returns (folder_url, att_count)."""
    try:
        drive_service = get_drive_service()
        dt_str        = datetime.now().strftime("%Y-%m-%d_%H%M")
        safe_name     = re.sub("[^a-zA-Z0-9 _-]", "", fields.get("customer_name") or "Unknown").strip().replace(" ", "_")
        folder_name   = f"{dt_str}_{safe_name}"
        folder_id, folder_url = create_drive_folder(drive_service, folder_name)

        # Upload Quote Template at root of quote folder
        excel_bytes = generate_quote_excel(email, fields, folder_url)
        upload_bytes_to_drive(drive_service, folder_id, "Quote_Template.xlsx", excel_bytes,
                              "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

        # Upload email attachments in dated incoming subfolder
        att_count   = 0
        attachments = email.get("attachments", [])
        if attachments:
            sub_folder_id = create_attachment_subfolder(drive_service, folder_id, "incoming")
            for att in attachments:
                if not att.get("attachment_id"):
                    continue
                try:
                    data = download_attachment(service, email["id"], att["attachment_id"])
                    upload_bytes_to_drive(drive_service, sub_folder_id, att["filename"], data, att.get("mime_type", ""))
                    att_count += 1
                except Exception as e:
                    st.warning(f"Could not upload {att['filename']}: {e}")

        return folder_url, att_count

    except Exception as e:
        st.error(f"Google Drive error: {e}")
        return "", 0


# ── Ignored Emails ────────────────────────────────────────────────────────────

def fetch_ignored_ids(supabase: Client) -> set:
    """Return set of message IDs that should be filtered out."""
    try:
        res = supabase.schema(get_schema()).table("ignored_emails").select("message_id").execute()
        return {row["message_id"] for row in res.data}
    except Exception:
        return set()


def fetch_blocked_patterns(supabase: Client) -> list:
    """Return list of blocked sender patterns (email or domain)."""
    try:
        res = supabase.schema(get_schema()).table("blocked_senders").select("pattern").execute()
        return [row["pattern"].lower().strip() for row in res.data]
    except Exception:
        return []


def is_blocked(sender: str, patterns: list) -> bool:
    """Check if sender matches any blocked pattern."""
    sender_lower = sender.lower()
    for pattern in patterns:
        if pattern.startswith("@"):
            # Domain block — match any email ending with this domain
            if pattern in sender_lower:
                return True
        else:
            # Exact email block
            if pattern in sender_lower:
                return True
    return False


def ignore_email(supabase: Client, email: dict):
    """Save a message ID to the ignored_emails table."""
    try:
        supabase.schema(get_schema()).table("ignored_emails").insert({
            "message_id": email["id"],
            "sender":     email["sender"],
            "subject":    email["subject"],
        }).execute()
    except Exception:
        pass


# ── Gmail Auth ─────────────────────────────────────────────────────────────────

@st.cache_resource
def get_gmail_service():
    gmail_secret = dict(st.secrets["gmail"])
    creds = Credentials(
        token=None,
        refresh_token=gmail_secret["refresh_token"],
        token_uri="https://oauth2.googleapis.com/token",
        client_id=gmail_secret["client_id"],
        client_secret=gmail_secret["client_secret"],
        scopes=SCOPES,
    )
    if not creds.valid:
        creds.refresh(Request())
    return build("gmail", "v1", credentials=creds)


# ── Gmail Helpers ──────────────────────────────────────────────────────────────

def decode_body(payload: dict) -> str:
    """Recursively extract plain text body from Gmail payload."""
    mime = payload.get("mimeType", "")
    if mime == "text/plain":
        data = payload.get("body", {}).get("data", "")
        if data:
            return base64.urlsafe_b64decode(data).decode("utf-8", errors="replace")
    if mime == "text/html" and not payload.get("parts"):
        data = payload.get("body", {}).get("data", "")
        if data:
            html = base64.urlsafe_b64decode(data).decode("utf-8", errors="replace")
            return re.sub(r"<[^>]+>", " ", html)
    for part in payload.get("parts", []):
        result = decode_body(part)
        if result:
            return result
    return ""


def get_attachments(payload: dict) -> list:
    """Return list of attachment metadata from Gmail payload."""
    attachments = []
    for part in payload.get("parts", []):
        filename = part.get("filename", "")
        if filename:
            attachments.append({
                "filename":    filename,
                "mime_type":   part.get("mimeType", ""),
                "attachment_id": part.get("body", {}).get("attachmentId", ""),
                "size":        part.get("body", {}).get("size", 0),
            })
        attachments.extend(get_attachments(part))
    return attachments


def download_attachment(service, message_id: str, attachment_id: str) -> bytes:
    """Download attachment bytes from Gmail."""
    result = service.users().messages().attachments().get(
        userId="me", messageId=message_id, id=attachment_id
    ).execute()
    data = result.get("data", "")
    return base64.urlsafe_b64decode(data)





def fetch_sent_emails(service) -> list:
    """Fetch sent emails from Gmail Sent folder."""
    import time

    def execute_with_retry(request, retries=3, delay=2):
        for attempt in range(retries):
            try:
                return request.execute()
            except Exception as e:
                if attempt < retries - 1 and ("Broken pipe" in str(e) or "Connection" in str(e) or "reset" in str(e).lower()):
                    time.sleep(delay)
                    continue
                raise

    result = execute_with_retry(
        service.users().messages().list(userId="me", labelIds=["SENT"], maxResults=50)
    )
    messages = result.get("messages", [])
    if not messages:
        return []

    emails = []
    for msg in messages:
        try:
            full    = execute_with_retry(service.users().messages().get(userId="me", id=msg["id"]))
            headers = {h["name"]: h["value"] for h in full["payload"]["headers"]}
            body    = decode_body(full["payload"])
            atts    = get_attachments(full["payload"])
            date_raw = headers.get("Date", "")
            try:
                from email.utils import parsedate_to_datetime
                dt = parsedate_to_datetime(date_raw)
                date_fmt = dt.strftime("%d %b, %I:%M %p")
            except Exception:
                date_fmt = date_raw[:16]
            emails.append({
                "id":          msg["id"],
                "thread_id":   full.get("threadId", ""),
                "subject":     headers.get("Subject", "(no subject)"),
                "sender":      headers.get("From", ""),
                "date":        date_fmt,
                "body":        body,
                "attachments": atts,
            })
        except Exception:
            continue
    return emails


def sync_sent_replies(service, supabase: Client) -> int:
    """
    Fetch sent emails, match to existing quote threads,
    append to conversation_log, upload any revised attachments.
    Returns count of threads updated.
    """
    # Get all known thread_ids from Supabase
    try:
        res = supabase.schema(get_schema()).table("quote_requests").select(
            "id, thread_id, conversation_log, reply_count, attachment_folder, "
            "customer_name, company_name, phone, sender_email, product_description, status"
        ).not_.is_("thread_id", "null").execute()
        known_threads = {row["thread_id"]: row for row in res.data}
    except Exception:
        return 0

    if not known_threads:
        return 0

    sent_emails = fetch_sent_emails(service)
    updated     = 0
    now         = datetime.now(timezone.utc).isoformat()

    for email in sent_emails:
        thread_id = email.get("thread_id", "")
        if thread_id not in known_threads:
            continue

        row = known_threads[thread_id]

        # Check if this message is already in conversation_log
        conv_log = row.get("conversation_log") or []
        existing_ids = {e.get("message_id") for e in conv_log}
        if email["id"] in existing_ids:
            continue  # already synced

        # Append sent message to conversation log
        conv_entry = {
            "message_id": email["id"],
            "sender":     email["sender"],
            "timestamp":  now,
            "subject":    email["subject"],
            "body":       email["body"][:2000],
            "type":       "sent",
        }
        conv_log.append(conv_entry)

        # Upload sent attachments in dated outgoing subfolder
        att_uploaded = 0
        folder_url   = row.get("attachment_folder", "")
        if email.get("attachments") and folder_url:
            try:
                drive_service = get_drive_service()
                folder_id     = folder_url.split("/")[-1]
                sub_folder_id = create_attachment_subfolder(drive_service, folder_id, "outgoing")
                for att in email["attachments"]:
                    if not att.get("attachment_id"):
                        continue
                    try:
                        data = download_attachment(service, email["id"], att["attachment_id"])
                        upload_bytes_to_drive(
                            drive_service, sub_folder_id,
                            att["filename"], data,
                            att.get("mime_type", "")
                        )
                        att_uploaded += 1
                    except Exception:
                        pass
            except Exception:
                pass

        # Update Supabase
        try:
            update_data = {
                "conversation_log": conv_log,
                "reply_count":      (row.get("reply_count") or 0) + 1,
                "last_reply_at":    now,
            }
            if att_uploaded:
                update_data["attachment_count"] = (row.get("attachment_count") or 0) + att_uploaded

            # Auto-set status to quoted if still new
            current_status = row.get("status", "new")
            if current_status == "new":
                update_data["status"] = "quoted"

            supabase.schema(get_schema()).table("quote_requests").update(update_data).eq("id", row["id"]).execute()

            # Append to Followup Tracker — only on first reply (reply_count was 0)
            if (row.get("reply_count") or 0) == 0:
                try:
                    drive_service = get_drive_service()
                    tracker_row = {
                        "id":                  row["id"],
                        "date":                datetime.now().strftime("%d-%b-%Y"),
                        "customer_name":       row.get("customer_name", ""),
                        "company_name":        row.get("company_name", ""),
                        "phone":               row.get("phone", ""),
                        "customer_email":      row.get("sender_email", ""),
                        "product_description": row.get("product_description", ""),
                        "status":              "quoted",
                        "attachment_folder":   row.get("attachment_folder", ""),
                    }
                    append_to_tracker(drive_service, tracker_row)
                except Exception as te:
                    st.warning(f"Tracker update failed: {te}")

            updated += 1
        except Exception as e:
            st.warning(f"Sync error: {e}")

    return updated


def fetch_unread_emails(service) -> list:
    """Fetch all unread emails from inbox with full details. Retries on connection errors."""
    import time

    def execute_with_retry(request, retries=3, delay=2):
        for attempt in range(retries):
            try:
                return request.execute()
            except Exception as e:
                if attempt < retries - 1 and ("Broken pipe" in str(e) or "Connection" in str(e) or "reset" in str(e).lower()):
                    time.sleep(delay)
                    continue
                raise

    result = execute_with_retry(
        service.users().messages().list(userId="me", labelIds=["INBOX", "UNREAD"], q="category:primary", maxResults=30)
    )
    messages = result.get("messages", [])
    if not messages:
        return []

    emails = []
    for msg in messages:
        full    = execute_with_retry(service.users().messages().get(userId="me", id=msg["id"]))
        headers = {h["name"]: h["value"] for h in full["payload"]["headers"]}
        body    = decode_body(full["payload"])
        atts    = get_attachments(full["payload"])

        # Format date
        date_raw = headers.get("Date", "")
        try:
            from email.utils import parsedate_to_datetime
            dt = parsedate_to_datetime(date_raw)
            date_fmt = dt.strftime("%d %b, %I:%M %p")
        except Exception:
            date_fmt = date_raw[:16]

        emails.append({
            "id":          msg["id"],
            "thread_id":   full.get("threadId", ""),
            "subject":     headers.get("Subject", "(no subject)"),
            "sender":      headers.get("From", ""),
            "date":        date_fmt,
            "body":        body,
            "preview":     body[:120].replace("\n", " ").strip(),
            "attachments": atts,
        })
    return emails


def mark_as_read(service, message_id: str):
    service.users().messages().modify(
        userId="me", id=message_id, body={"removeLabelIds": ["UNREAD"]}
    ).execute()


# ── Claude Extraction ──────────────────────────────────────────────────────────

def extract_quote_fields(email: dict) -> dict | None:
    try:
        api_key = st.secrets["ANTHROPIC_API_KEY"]
    except Exception as e:
        st.error(f"Missing ANTHROPIC_API_KEY in secrets: {e}")
        return None
    client = anthropic.Anthropic(api_key=api_key)
    prompt = EXTRACTION_PROMPT.format(
        app_name=get_app_name(),
        subject=email["subject"],
        sender=email["sender"],
        body=email["body"][:4000],
    )
    try:
        message = client.messages.create(
            model="claude-haiku-4-5-20251001",
            max_tokens=1024,
            messages=[{"role": "user", "content": prompt}],
        )
        raw = message.content[0].text.strip()
        raw = re.sub(r"^```(?:json)?", "", raw).strip()
        raw = re.sub(r"```$", "", raw).strip()
        return json.loads(raw)
    except Exception:
        return None


# ── Supabase Insert / Update ───────────────────────────────────────────────────

def upsert_quote(supabase: Client, service, email: dict, fields: dict) -> tuple[bool, str]:
    """
    Insert new quote or append to existing thread.
    Returns (success, action) where action is 'inserted' or 'updated'.
    """
    thread_id = email.get("thread_id", "")
    now       = datetime.now(timezone.utc).isoformat()

    # Build conversation entry for this message
    conv_entry = {
        "message_id": email.get("id", ""),
        "sender":    email["sender"],
        "timestamp": now,
        "subject":   email["subject"],
        "body":      email["body"][:2000],
        "type":      "received",
    }

    # Check if thread already exists
    existing = None
    if thread_id:
        try:
            res = supabase.schema(get_schema()).table("quote_requests").select("id, conversation_log, reply_count").eq("thread_id", thread_id).execute()
            if res.data:
                existing = res.data[0]
        except Exception:
            pass

    if existing:
        # Thread exists — append to conversation log
        conv_log   = existing.get("conversation_log") or []
        conv_log.append(conv_entry)
        reply_count = (existing.get("reply_count") or 0) + 1
        try:
            supabase.schema(get_schema()).table("quote_requests").update({
                "conversation_log": conv_log,
                "reply_count":      reply_count,
                "last_reply_at":    now,
            }).eq("id", existing["id"]).execute()
            return True, "updated"
        except Exception as e:
            return False, str(e)
    else:
        # New thread — save to Drive (attachments + Excel) and insert
        folder_path, att_count = save_to_drive(service, email, fields)

        row = {
            "thread_id":           thread_id or None,
            "customer_name":       fields.get("customer_name"),
            "customer_email":      fields.get("customer_email"),
            "company_name":        fields.get("company_name"),
            "phone":               fields.get("phone"),
            "product_description": fields.get("product_description"),
            "quantity":            str(fields.get("quantity")) if fields.get("quantity") else None,
            "unit":                fields.get("unit"),
            "deadline":            fields.get("deadline"),
            "location":            fields.get("location"),
            "urgency_level":       fields.get("urgency_level", "medium"),
            "needs_review":        bool(fields.get("needs_review", False)),
            "notes":               fields.get("notes"),
            "raw_email_subject":   email["subject"],
            "raw_email_body":      email["body"][:5000],
            "sender_email":        email["sender"],
            "status":              "new",
            "conversation_log":    [conv_entry],
            "reply_count":         0,
            "last_reply_at":       now,
            "attachment_folder":   folder_path or None,
            "attachment_count":    att_count,
        }
        try:
            supabase.schema(get_schema()).table("quote_requests").insert(row).execute()
            return True, "inserted"
        except Exception as e:
            return False, str(e)


# ── Streamlit UI ───────────────────────────────────────────────────────────────

st.set_page_config(
    page_title="Quote Extractor — Carob Technologies",
    page_icon="📬",
    layout="wide",
)

st.title("📬 QuoteDesk")
st.caption(f"Mail to Order · {get_app_name()} · Powered by Carob Technologies")


# Validate email provider
_supported_providers = ["gmail"]
_provider = get_email_provider()
if _provider not in _supported_providers:
    st.error(
        f"EMAIL_PROVIDER = '{_provider}' is not supported yet. "
        f"Supported providers: {', '.join(_supported_providers)}. "
        f"Please update your secrets.toml."
    )
    st.stop()


tab_inbox, tab_quotes, tab_analytics, tab_followup, tab_settings = st.tabs(["📬 Inbox", "📋 Quote Requests", "📊 Analytics", "📊 Track Status", "⚙️ Settings"])


# ── Tab 1: Inbox ───────────────────────────────────────────────────────────────

with tab_inbox:

    # Session state
    if "emails"   not in st.session_state: st.session_state.emails   = []
    if "selected" not in st.session_state: st.session_state.selected = set()
    if "log"      not in st.session_state: st.session_state.log      = ""

    col_fetch, col_selall, col_clear, col_count = st.columns([2, 1, 1, 3])

    with col_fetch:
        if st.button("📬 Fetch unread emails", type="primary", use_container_width=True):
            with st.spinner("Connecting to Gmail... (may retry on connection errors)"):
                try:
                    get_gmail_service.clear()  # force fresh connection
                    service     = get_gmail_service()
                    all_emails  = fetch_unread_emails(service)
                    ignored_ids      = fetch_ignored_ids(get_supabase())
                    blocked_patterns = fetch_blocked_patterns(get_supabase())
                    st.session_state.emails   = [
                        e for e in all_emails
                        if e["id"] not in ignored_ids
                        and not is_blocked(e["sender"], blocked_patterns)
                    ]
                    st.session_state.selected = set()
                    st.session_state.log      = ""
                    ignored_count = len(all_emails) - len(st.session_state.emails)
                    if ignored_count:
                        st.caption(f"ℹ️ {ignored_count} ignored email(s) filtered out.")
                    # Sync sent replies in background
                    sent_synced = sync_sent_replies(service, get_supabase())
                    if sent_synced:
                        st.caption(f"🔄 {sent_synced} sent reply(s) synced to conversation log.")
                except Exception as e:
                    st.error(f"Gmail error: {e}")

    with col_selall:
        if st.button("☑ Select all", use_container_width=True):
            st.session_state.selected = {e["id"] for e in st.session_state.emails}
            st.rerun()

    with col_clear:
        if st.button("☐ Clear", use_container_width=True):
            st.session_state.selected = set()
            st.rerun()

    with col_count:
        total    = len(st.session_state.emails)
        selected = len(st.session_state.selected)
        if total:
            st.info(f"{total} unread email(s) fetched · {selected} selected", icon="📨")

    st.info(
        "Unselected emails stay **unread** in Gmail. "
        "Only selected emails are extracted and saved to Supabase.",
        icon="ℹ️",
    )

    # Email list
    if not st.session_state.emails:
        st.markdown("Click **Fetch unread emails** to load your inbox.")
    else:
        for email in st.session_state.emails:
            is_checked = email["id"] in st.session_state.selected
            has_att    = len(email.get("attachments", [])) > 0
            att_label  = f" 📎 {len(email['attachments'])}" if has_att else ""

            col_cb, col_body = st.columns([0.5, 11])

            with col_cb:
                st.write("")
                checked = st.checkbox(
                    "Select",
                    value=is_checked,
                    key=f"chk_{email['id']}",
                    label_visibility="collapsed",
                )
                if checked:
                    st.session_state.selected.add(email["id"])
                else:
                    st.session_state.selected.discard(email["id"])

            with col_body:
                with st.expander(
                    f"**{email['sender']}** · {email['subject']}{att_label} · *{email['date']}*"
                ):
                    st.write(email["body"] or "(no body)")
                    if has_att:
                        st.caption(f"📎 Attachments: {', '.join(a['filename'] for a in email['attachments'])}")
                    if st.button("🚫 Ignore this email", key=f"ign_{email['id']}", help="Hide this email permanently — it will never appear in this app again"):
                        ignore_email(get_supabase(), email)
                        st.session_state.emails   = [e for e in st.session_state.emails if e["id"] != email["id"]]
                        st.session_state.selected.discard(email["id"])
                        st.rerun()

        st.divider()

        # Extract bar
        n_selected = len(st.session_state.selected)
        col_ex, col_status = st.columns([2, 4])

        with col_ex:
            extract_clicked = st.button(
                f"▶ Extract {n_selected} selected email(s) → Supabase",
                type="primary",
                disabled=(n_selected == 0),
                use_container_width=True,
            )

        if extract_clicked and n_selected > 0:
            supabase = get_supabase()
            service  = get_gmail_service()
            log_area = st.empty()
            log_buf  = []

            def log(text):
                log_buf.append(text)
                log_area.code("".join(log_buf), language=None)

            inserted = updated = failed = 0

            for email in st.session_state.emails:
                if email["id"] not in st.session_state.selected:
                    continue

                log(f"📋 Processing: {email['subject']}\n")
                log(f"   From: {email['sender']}\n")

                fields = extract_quote_fields(email)
                if not fields:
                    log("   ❌ Claude extraction failed — left unread.\n\n")
                    failed += 1
                    continue

                log(f"   Customer : {fields.get('customer_name') or '—'}\n")
                log(f"   Product  : {fields.get('product_description') or '—'}\n")
                log(f"   Quantity : {fields.get('quantity') or '—'} {fields.get('unit') or ''}\n")
                log(f"   Deadline : {fields.get('deadline') or '—'}\n")
                log(f"   Urgency  : {fields.get('urgency_level') or '—'}\n")
                log(f"   Review?  : {'⚠️ Yes' if fields.get('needs_review') else '✅ No'}\n")

                ok, action = upsert_quote(supabase, service, email, fields)
                if ok:
                    mark_as_read(service, email["id"])
                    if action == "inserted":
                        log("   ✅ New quote saved to Supabase.\n\n")
                        inserted += 1
                    else:
                        log("   🔄 Existing thread updated in Supabase.\n\n")
                        updated += 1
                else:
                    log(f"   ❌ Supabase error: {action} — left unread.\n\n")
                    failed += 1

            log(f"─────────────────────────────────────\n")
            log(f"Done. {inserted} new · {updated} thread updates · {failed} failed.\n")

            # Refresh inbox
            st.session_state.emails   = fetch_unread_emails(service)
            st.session_state.selected = set()

            if failed == 0:
                st.success(f"✅ {inserted} saved · {updated} updated. Switch to Quote Requests tab.")
            else:
                st.warning(f"{inserted} saved · {updated} updated · {failed} failed (still unread in Gmail).")


# ── Tab 2: Quote Requests ──────────────────────────────────────────────────────

with tab_quotes:
    supabase = get_supabase()

    col1, col2, col3 = st.columns([2, 2, 1])
    with col1:
        status_filter  = st.selectbox("Status", ["All"] + STATUS_OPTIONS)
    with col2:
        urgency_filter = st.selectbox("Urgency", ["All", "high", "medium", "low"])
    with col3:
        st.write("")
        st.button("🔄 Refresh", use_container_width=True, key="refresh_quotes")

    try:
        query = supabase.schema(get_schema()).table("quote_requests").select("*").order("created_at", desc=True)
        if status_filter  != "All": query = query.eq("status", status_filter)
        if urgency_filter != "All": query = query.eq("urgency_level", urgency_filter)
        rows = query.execute().data
    except Exception as e:
        st.error(f"Could not load from Supabase: {e}")
        rows = []

    if not rows:
        st.info("No quote requests found.")
    else:
        st.caption(f"{len(rows)} record(s)")

        for row in rows:
            created     = row.get("created_at", "")[:16].replace("T", " ")
            status      = row.get("status", "new")
            urgency     = row.get("urgency_level", "medium")
            review      = row.get("needs_review", False)
            reply_count = row.get("reply_count", 0)
            s_icon      = STATUS_COLORS.get(status, "⚪")
            u_icon      = URGENCY_ICONS.get(urgency, "⚪")

            label = (
                f"{s_icon} {row.get('customer_name') or 'Unknown'}  —  "
                f"{(row.get('product_description') or '')[:50]}  "
                f"{'⚠️' if review else ''}  "
                f"{'💬 ' + str(reply_count + 1) if reply_count else ''}  "
                f"|  {created}"
            )

            with st.expander(label):
                left, right = st.columns(2)

                with left:
                    st.markdown("**Customer**")
                    st.write(f"Name    : {row.get('customer_name') or '—'}")
                    st.write(f"Email   : {row.get('customer_email') or '—'}")
                    st.write(f"Company : {row.get('company_name') or '—'}")
                    st.write(f"Phone   : {row.get('phone') or '—'}")

                    st.markdown("**Request**")
                    st.write(f"Product  : {row.get('product_description') or '—'}")
                    st.write(f"Quantity : {row.get('quantity') or '—'} {row.get('unit') or ''}")
                    st.write(f"Deadline : {row.get('deadline') or '—'}")
                    st.write(f"Location : {row.get('location') or '—'}")

                    if row.get("attachment_folder"):
                        st.markdown("**Attachments**")
                        if row.get("attachment_folder"):
                            st.markdown(f"[📁 Open in Google Drive]({row['attachment_folder']})")
                        st.write(f"📎 {row.get('attachment_count', 0)} attachment(s) + Quote_Template.xlsx")

                with right:
                    st.markdown("**AI Assessment**")
                    st.write(f"Urgency      : {u_icon} {urgency}")
                    st.write(f"Needs review : {'⚠️ Yes' if review else '✅ No'}")
                    if row.get("notes"):
                        st.write(f"Notes : {row['notes']}")
                    if row.get("last_reply_at"):
                        st.write(f"Last reply   : {row['last_reply_at'][:16].replace('T',' ')}")

                    st.markdown("**Update Status**")
                    new_status = st.selectbox(
                        "Status",
                        STATUS_OPTIONS,
                        index=STATUS_OPTIONS.index(status),
                        key=f"status_{row['id']}",
                        label_visibility="collapsed",
                    )
                    if new_status != status:
                        try:
                            supabase.schema(get_schema()).table("quote_requests").update(
                                {"status": new_status}
                            ).eq("id", row["id"]).execute()
                            # Remove from tracker if closed
                            if new_status in ["won", "lost", "not_relevant", "not_feasible"]:
                                try:
                                    remove_from_tracker(get_drive_service(), row["id"])
                                except Exception:
                                    pass
                            st.success(f"Updated to **{new_status}**")
                            st.rerun()
                        except Exception as e:
                            st.error(f"Update failed: {e}")

                # Original email
                with st.expander("📧 Original email"):
                    st.write(f"**Subject:** {row.get('raw_email_subject') or '—'}")
                    st.write(f"**From:** {row.get('sender_email') or '—'}")
                    st.text(row.get("raw_email_body") or "—")

                # Merge feature
                with st.expander("🔗 Merge into another quote"):
                    st.caption("Use this if this email is a follow-up from a different person in the same organisation, related to an existing quote.")
                    try:
                        all_quotes = supabase.schema(get_schema()).table("quote_requests").select(
                            "id, customer_name, customer_email, product_description, created_at"
                        ).neq("id", row["id"]).order("created_at", desc=True).execute().data

                        if not all_quotes:
                            st.info("No other quotes to merge into.")
                        else:
                            merge_options = {
                                f"{q.get('customer_name') or q.get('customer_email') or 'Unknown'} — {(q.get('product_description') or '')[:40]} ({q.get('created_at', '')[:10]})": q["id"]
                                for q in all_quotes
                            }
                            selected_label = st.selectbox(
                                "Select quote to merge into:",
                                list(merge_options.keys()),
                                key=f"merge_select_{row['id']}"
                            )
                            st.warning(f"This will delete the current record and append its conversation to the selected quote.")
                            if st.button("🔗 Confirm Merge", key=f"merge_btn_{row['id']}", type="primary"):
                                target_id = merge_options[selected_label]
                                service   = get_gmail_service()
                                ok = merge_quotes(supabase, row["id"], target_id, service)
                                if ok:
                                    st.success("Merged successfully!")
                                    st.rerun()
                    except Exception as e:
                        st.error(f"Could not load quotes: {e}")

                # Conversation thread
                conv_log = row.get("conversation_log") or []
                if conv_log:
                    st.markdown("---")
                    st.markdown(f"**💬 Conversation ({len(conv_log)} message(s))**")
                    for entry in conv_log:
                        sender    = entry.get("sender", "")
                        timestamp = entry.get("timestamp", "")[:16].replace("T", " ")
                        body      = entry.get("body", "")
                        subject   = entry.get("subject", "")
                        is_me     = "carobashok" in sender.lower()
                        align     = "🏢" if is_me else "👤"
                        name      = "You" if is_me else sender.split("<")[0].strip()

                        st.markdown(f"{align} **{name}** · *{timestamp}* · _{subject}_")
                        st.text(body[:500] + ("..." if len(body) > 500 else ""))
                        st.markdown("---")


# ── Tab 3: Analytics ───────────────────────────────────────────────────────────

with tab_analytics:
    supabase = get_supabase()

    # Date filter
    col_f1, col_f2 = st.columns([3, 1])
    with col_f1:
        st.markdown("### 📊 Quote Analytics")
    with col_f2:
        period = st.selectbox("Period", ["Last 30 days", "All time"], label_visibility="collapsed")

    # Fetch all records
    try:
        rows = supabase.schema(get_schema()).table("quote_requests").select("*").order("created_at", desc=True).execute().data
    except Exception as e:
        st.error(f"Could not load data: {e}")
        rows = []

    if not rows:
        st.info("No quote requests found.")
    else:
        from datetime import timedelta
        import pandas as pd

        # Convert to DataFrame
        df = pd.DataFrame(rows)
        df["created_at"] = pd.to_datetime(df["created_at"], utc=True)

        # Apply period filter for KPIs
        if period == "Last 30 days":
            cutoff = datetime.now(timezone.utc) - timedelta(days=30)
            df_filtered = df[df["created_at"] >= cutoff]
        else:
            df_filtered = df

        # ── KPI Cards ─────────────────────────────────────────────────────────

        total   = len(df_filtered)
        won     = len(df_filtered[df_filtered["status"] == "won"])
        lost    = len(df_filtered[df_filtered["status"] == "lost"])
        pending = len(df_filtered[df_filtered["status"].isin(["new", "quoted"])])
        not_relevant = len(df_filtered[df_filtered["status"] == "not_relevant"])
        not_feasible = len(df_filtered[df_filtered["status"] == "not_feasible"])

        # Avg response time — from conversation_log
        response_times = []
        for _, row in df_filtered.iterrows():
            conv_log = row.get("conversation_log") or []
            if not conv_log:
                continue
            received_ts = None
            sent_ts     = None
            for entry in conv_log:
                if entry.get("type") in ("received", None, "") and received_ts is None:
                    received_ts = entry.get("timestamp")
                if entry.get("type") == "sent" and sent_ts is None:
                    sent_ts = entry.get("timestamp")
            if received_ts and sent_ts:
                try:
                    t1 = pd.to_datetime(received_ts, utc=True)
                    t2 = pd.to_datetime(sent_ts, utc=True)
                    diff_hours = (t2 - t1).total_seconds() / 3600
                    if diff_hours >= 0:
                        response_times.append(diff_hours)
                except Exception:
                    pass

        avg_response = f"{sum(response_times)/len(response_times):.1f} hrs" if response_times else "—"

        # Display KPI cards
        k1, k2, k3, k4, k5, k6, k7 = st.columns(7)
        k1.metric("📥 Total Quotes",   total)
        k2.metric("🟢 Won",            won)
        k3.metric("🔴 Lost",           lost)
        k4.metric("🔵 Pending",        pending)
        k5.metric("⚫ Not Relevant",   not_relevant)
        k6.metric("🟤 Not Feasible",   not_feasible)
        k7.metric("⏱ Avg Response",    avg_response)

        st.divider()

        # ── Charts ────────────────────────────────────────────────────────────

        col_chart1, col_chart2 = st.columns(2)

        with col_chart1:
            st.markdown("**📈 Quote Volume — Last 30 days**")
            cutoff30 = datetime.now(timezone.utc) - timedelta(days=30)
            df30 = df[df["created_at"] >= cutoff30].copy()
            if not df30.empty:
                df30["date"] = df30["created_at"].dt.date
                vol = df30.groupby("date").size().reset_index(name="count")
                vol["date"] = vol["date"].astype(str)
                st.bar_chart(vol.set_index("date")["count"])
            else:
                st.info("No data for last 30 days.")

        with col_chart2:
            st.markdown("**🔴 Urgency Distribution**")
            if not df_filtered.empty:
                urgency_counts = df_filtered["urgency_level"].value_counts()
                st.bar_chart(urgency_counts)
            else:
                st.info("No data.")

        st.divider()

        # ── Conversion Funnel ─────────────────────────────────────────────────

        st.markdown("**🔄 Conversion Pipeline**")
        funnel_data = {
            "New":          len(df_filtered[df_filtered["status"] == "new"]),
            "Quoted":       len(df_filtered[df_filtered["status"] == "quoted"]),
            "Won":          won,
            "Lost":         lost,
            "Not Relevant": not_relevant,
            "Not Feasible": not_feasible,
        }
        funnel_df = pd.DataFrame(list(funnel_data.items()), columns=["Stage", "Count"])
        st.bar_chart(funnel_df.set_index("Stage"))

        st.divider()

        # ── Detail Table — Last 30 days ───────────────────────────────────────

        st.markdown("**📋 Quote Detail — Last 30 days**")
        cutoff30 = datetime.now(timezone.utc) - timedelta(days=30)
        df_detail = df[df["created_at"] >= cutoff30].copy()

        if df_detail.empty:
            st.info("No quotes in the last 30 days.")
        else:
            table_rows = []
            for _, row in df_detail.iterrows():
                conv_log    = row.get("conversation_log") or []
                received_ts = None
                sent_ts     = None
                for entry in conv_log:
                    if entry.get("type") in ("received", None, "") and received_ts is None:
                        received_ts = entry.get("timestamp")
                    if entry.get("type") == "sent" and sent_ts is None:
                        sent_ts = entry.get("timestamp")

                # Response time
                if received_ts and sent_ts:
                    try:
                        t1 = pd.to_datetime(received_ts, utc=True)
                        t2 = pd.to_datetime(sent_ts, utc=True)
                        diff = (t2 - t1).total_seconds() / 3600
                        resp_str = f"{diff:.1f} hrs" if diff >= 0 else "—"
                    except Exception:
                        resp_str = "—"
                else:
                    resp_str = "Not replied yet"

                table_rows.append({
                    "Customer Email":   row.get("customer_email") or "—",
                    "Product/Service":  (row.get("product_description") or "—")[:50],
                    "Date of Request":  row.get("created_at").strftime("%d %b %Y") if pd.notnull(row.get("created_at")) else "—",
                    "Date of Reply":    pd.to_datetime(sent_ts, utc=True).strftime("%d %b %Y") if sent_ts else "—",
                    "Response Time":    resp_str,
                    "Status":           row.get("status", "—"),
                })

            detail_df = pd.DataFrame(table_rows)
            st.dataframe(detail_df, use_container_width=True, hide_index=True)


# ── Tab 4: Follow Up ───────────────────────────────────────────────────────────

with tab_followup:
    supabase = get_supabase()

    st.markdown("### 📊 Track Status")

    # Filters
    col_f1, col_f2, col_f3 = st.columns([2, 2, 1])
    with col_f1:
        fu_status_filter = st.selectbox(
            "Follow Up Status",
            ["All", "pending", "contacted", "interested"],
            key="fu_status_filter"
        )
    with col_f2:
        quote_status_filter = st.selectbox(
            "Quote Status",
            ["All"] + STATUS_OPTIONS,
            key="fu_quote_filter"
        )
    with col_f3:
        st.write("")
        st.button("🔄 Refresh", use_container_width=True, key="fu_refresh")

    # Fetch records
    try:
        query = supabase.schema(get_schema()).table("quote_requests").select("*").order("followup_date", desc=False)
        if fu_status_filter != "All":
            query = query.eq("followup_status", fu_status_filter)
        if quote_status_filter != "All":
            query = query.eq("status", quote_status_filter)
        fu_rows = query.execute().data
    except Exception as e:
        st.error(f"Could not load follow-up data: {e}")
        fu_rows = []

    FOLLOWUP_COLORS = {
        "pending":    "🟡",
        "contacted":  "🔵",
        "interested": "🟢",
    }

    if not fu_rows:
        st.info("No follow-up records found.")
    else:
        # Summary counts
        c1, c2, c3 = st.columns(3)
        c1.metric("🟡 Pending",    sum(1 for r in fu_rows if r.get("followup_status") == "pending"))
        c2.metric("🔵 Contacted",  sum(1 for r in fu_rows if r.get("followup_status") == "contacted"))
        c3.metric("🟢 Interested", sum(1 for r in fu_rows if r.get("followup_status") == "interested"))

        st.divider()

        # Due today / overdue highlight
        from datetime import date
        today = date.today()

        for row in fu_rows:
            fu_status   = row.get("followup_status", "pending")
            fu_date     = row.get("followup_date")
            fu_icon     = FOLLOWUP_COLORS.get(fu_status, "⚪")
            q_status    = row.get("status", "new")
            q_icon      = STATUS_COLORS.get(q_status, "⚪")
            customer    = row.get("customer_email") or row.get("customer_name") or "Unknown"
            product     = (row.get("product_description") or "")[:45]
            notes_log   = row.get("followup_notes") or []

            # Date display
            if fu_date:
                fu_date_obj = date.fromisoformat(fu_date) if isinstance(fu_date, str) else fu_date
                if fu_date_obj < today:
                    date_label = f"⚠️ Overdue: {fu_date}"
                elif fu_date_obj == today:
                    date_label = f"🔴 Due Today: {fu_date}"
                else:
                    date_label = f"📅 {fu_date}"
            else:
                date_label = "No date set"

            label = f"{fu_icon} {customer} — {product} | {q_icon} {q_status} | {date_label}"

            with st.expander(label):
                left, right = st.columns(2)

                with left:
                    st.markdown("**Customer**")
                    st.write(f"Name    : {row.get('customer_name') or '—'}")
                    st.write(f"Email   : {row.get('customer_email') or '—'}")
                    st.write(f"Company : {row.get('company_name') or '—'}")
                    st.write(f"Phone   : {row.get('phone') or '—'}")
                    st.markdown("**Quote**")
                    st.write(f"Product  : {row.get('product_description') or '—'}")
                    st.write(f"Status   : {q_icon} {q_status}")
                    if row.get("attachment_folder"):
                        st.markdown(f"[📁 Open in Drive]({row['attachment_folder']})")

                with right:
                    st.markdown("**Update Follow Up**")

                    new_fu_status = st.selectbox(
                        "Follow Up Status",
                        ["pending", "contacted", "interested"],
                        index=["pending", "contacted", "interested"].index(fu_status) if fu_status in ["pending", "contacted", "interested"] else 0,
                        key=f"fu_status_{row['id']}",
                    )

                    new_fu_date = st.date_input(
                        "Next Follow Up Date",
                        value=date.fromisoformat(fu_date) if fu_date and isinstance(fu_date, str) else (fu_date if fu_date else today),
                        key=f"fu_date_{row['id']}",
                    )

                    new_note = st.text_input(
                        "Add note (call/email summary)",
                        placeholder="e.g. Called — will revert by Friday",
                        key=f"fu_note_{row['id']}",
                    )

                    if st.button("💾 Save", key=f"fu_save_{row['id']}", type="primary"):
                        update_data = {
                            "followup_status": new_fu_status,
                            "followup_date":   new_fu_date.isoformat(),
                        }
                        if new_note.strip():
                            notes_log.append({
                                "timestamp": datetime.now(timezone.utc).isoformat(),
                                "note":      new_note.strip(),
                                "status":    new_fu_status,
                            })
                            update_data["followup_notes"] = notes_log
                        try:
                            supabase.schema(get_schema()).table("quote_requests").update(update_data).eq("id", row["id"]).execute()
                            st.success("Saved!")
                            st.rerun()
                        except Exception as e:
                            st.error(f"Save failed: {e}")

                # Follow up history
                if notes_log:
                    st.markdown("---")
                    st.markdown(f"**📋 Follow Up History ({len(notes_log)} entries)**")
                    for entry in reversed(notes_log):
                        ts     = entry.get("timestamp", "")[:16].replace("T", " ")
                        status = entry.get("status", "")
                        note   = entry.get("note", "")
                        icon   = FOLLOWUP_COLORS.get(status, "⚪")
                        st.markdown(f"{icon} **{ts}** — {note}")


# ── Tab 5: Settings ────────────────────────────────────────────────────────────

with tab_settings:
    supabase = get_supabase()

    st.markdown("### ⚙️ Settings")

    # ── Blocked Senders ────────────────────────────────────────────────────────
    st.markdown("#### 🚫 Blocked Senders")
    st.caption(
        "Emails from these addresses or domains will never appear in the Inbox. "
        "Use @domain.com to block an entire domain (e.g. @linkedin.com)."
    )

    # Fetch current blocked list
    try:
        blocked = supabase.schema(get_schema()).table("blocked_senders").select("*").order("created_at", desc=False).execute().data
    except Exception as e:
        st.error(f"Could not load blocked senders: {e}")
        blocked = []

    # Display existing blocked senders
    if not blocked:
        st.info("No blocked senders yet.")
    else:
        for b in blocked:
            col_pattern, col_note, col_del = st.columns([3, 4, 1])
            with col_pattern:
                st.code(b.get("pattern", ""), language=None)
            with col_note:
                st.caption(b.get("note") or "—")
            with col_del:
                if st.button("✕", key=f"del_blocked_{b['id']}", help="Remove this block"):
                    try:
                        supabase.schema(get_schema()).table("blocked_senders").delete().eq("id", b["id"]).execute()
                        st.rerun()
                    except Exception as e:
                        st.error(f"Could not remove: {e}")

    st.divider()

    # Add new blocked sender
    st.markdown("**Add blocked sender or domain:**")
    col_add1, col_add2, col_add3 = st.columns([3, 4, 1])
    with col_add1:
        new_pattern = st.text_input(
            "Email or domain",
            placeholder="e.g. noreply@linkedin.com or @newsletter.com",
            label_visibility="collapsed",
            key="new_block_pattern"
        )
    with col_add2:
        new_note = st.text_input(
            "Note (optional)",
            placeholder="e.g. LinkedIn notifications",
            label_visibility="collapsed",
            key="new_block_note"
        )
    with col_add3:
        if st.button("➕ Add", type="primary", use_container_width=True):
            if new_pattern.strip():
                try:
                    # Support comma-separated multiple entries
                    patterns = [p.strip().lower() for p in new_pattern.split(",") if p.strip()]
                    added = []
                    for p in patterns:
                        try:
                            supabase.schema(get_schema()).table("blocked_senders").insert({
                                "pattern": p,
                                "note":    new_note.strip() or None,
                            }).execute()
                            added.append(p)
                        except Exception:
                            pass  # skip duplicates silently
                    if added:
                        st.success(f"Blocked {len(added)} sender(s): {', '.join(added)}")
                        st.rerun()
                except Exception as e:
                    st.error(f"Could not add: {e}")
            else:
                st.warning("Please enter an email or domain to block.")

    st.divider()

    # ── Ignored Emails ─────────────────────────────────────────────────────────
    st.markdown("#### 🙈 Ignored Emails")
    st.caption("Individual emails you have ignored. These will never reappear in the Inbox.")

    try:
        ignored = supabase.schema(get_schema()).table("ignored_emails").select("*").order("ignored_at", desc=True).execute().data
    except Exception as e:
        st.error(f"Could not load ignored emails: {e}")
        ignored = []

    if not ignored:
        st.info("No ignored emails.")
    else:
        st.caption(f"{len(ignored)} ignored email(s)")
        for ig in ignored:
            col_i1, col_i2, col_i3 = st.columns([3, 4, 1])
            with col_i1:
                st.caption(ig.get("sender", "—"))
            with col_i2:
                st.caption(ig.get("subject", "—")[:60])
            with col_i3:
                if st.button("↩️", key=f"unignore_{ig['id']}", help="Restore this email"):
                    try:
                        supabase.schema(get_schema()).table("ignored_emails").delete().eq("id", ig["id"]).execute()
                        st.success("Restored — will appear on next Fetch.")
                        st.rerun()
                    except Exception as e:
                        st.error(f"Could not restore: {e}")
